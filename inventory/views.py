from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

from products.models import Product
from .models import InventoryRecord
from notifications.models import *

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from django.utils.timezone import make_naive

from .models import Product, InventoryRecord
from sales.models import Sale

@login_required(login_url='login')
def export_inventory_to_excel(request):
    """Exporte les données d'inventaire vers un fichier Excel."""
    # Création d'un classeur et ajout de feuilles
    wb = Workbook()

    # Définition des styles
    header_fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")  # Fond bleu
    header_font = Font(color="FFFFFF", bold=True)  # Texte blanc en gras
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    alignment = Alignment(horizontal="left", vertical="center")

    # Feuille 1 : Tout l'inventaire
    ws_all = wb.active
    ws_all.title = "Tout l'inventaire"
    ws_all.append(['Nom', 'Marque', 'Catégorie', 'Quantité en stock'])

    # Appliquer les styles de l'en-tête
    for col in range(1, 5):
        cell = ws_all.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = alignment

    # Ajouter les données et appliquer les styles
    products = Product.objects.all()
    for product in products:
        ws_all.append([product.name, product.get_brand_display(), product.get_category_display(), product.quantity])

    for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

    # Définir les largeurs des colonnes
    ws_all.column_dimensions['A'].width = 30  # Nom
    ws_all.column_dimensions['B'].width = 20  # Marque
    ws_all.column_dimensions['C'].width = 20  # Catégorie
    ws_all.column_dimensions['D'].width = 15  # Quantité

    # Feuille 2 : Produits enregistrés (Historique des inventaires)
    ws_registered = wb.create_sheet("Produits enregistrés")
    ws_registered.append(['Nom du produit', 'Marque', 'Catégorie', 'Quantité enregistrée', 'Date', 'Enregistré par'])

    # Appliquer les styles de l'en-tête
    for col in range(1, 7):
        cell = ws_registered.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = alignment

    # Ajouter les données et appliquer les styles
    inventory_records = InventoryRecord.objects.all()
    for record in inventory_records:
        naive_date = make_naive(record.date)  # Convertir la date en temps local
        ws_registered.append([
            record.product.name,
            record.product.get_brand_display(),
            record.product.get_category_display(),
            record.quantity_recorded,
            naive_date,
            record.done_by.name
        ])

    for row in ws_registered.iter_rows(min_row=2, max_row=ws_registered.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

    # Définir les largeurs des colonnes
    ws_registered.column_dimensions['A'].width = 30  # Nom du produit
    ws_registered.column_dimensions['B'].width = 20  # Marque
    ws_registered.column_dimensions['C'].width = 20  # Catégorie
    ws_registered.column_dimensions['D'].width = 15  # Quantité enregistrée
    ws_registered.column_dimensions['E'].width = 20  # Date
    ws_registered.column_dimensions['F'].width = 20  # Enregistré par

    # Feuille 3 : Produits vendus
    ws_sold = wb.create_sheet("Produits vendus")
    ws_sold.append(['Nom du produit', 'Marque', 'Catégorie', 'Quantité vendue', 'Prix unitaire', 'Prix total', 'Vendu par'])

    # Appliquer les styles de l'en-tête
    for col in range(1, 8):
        cell = ws_sold.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = alignment

    # Ajouter les données et appliquer les styles
    sales = Sale.objects.all()
    for sale in sales:
        ws_sold.append([
            sale.product.name,
            sale.product.get_brand_display(),
            sale.product.get_category_display(),
            sale.quantity,
            sale.product.unit_selling_price,
            sale.total_price,
            sale.done_by.name
        ])

    for row in ws_sold.iter_rows(min_row=2, max_row=ws_sold.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = alignment

    # Définir les largeurs des colonnes
    ws_sold.column_dimensions['A'].width = 30  # Nom du produit
    ws_sold.column_dimensions['B'].width = 20  # Marque
    ws_sold.column_dimensions['C'].width = 20  # Catégorie
    ws_sold.column_dimensions['D'].width = 15  # Quantité vendue
    ws_sold.column_dimensions['E'].width = 15  # Prix unitaire
    ws_sold.column_dimensions['F'].width = 15  # Prix total
    ws_sold.column_dimensions['G'].width = 20  # Vendu par

    # Définir les hauteurs des lignes
    for sheet in wb:
        for row in sheet.iter_rows():
            sheet.row_dimensions[row[0].row].height = 20  # Hauteur de ligne fixée à 20

    # Sauvegarder le classeur dans une réponse HTTP
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="rapport_inventaire.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='login')
def inventory_view(request):
    """Affiche la vue de l'inventaire avec les produits, les enregistrements et les ventes."""
    products = Product.objects.all()
    context = {
        'products': products,
        'inventory_records': InventoryRecord.objects.all(),
        'sales': Sale.objects.all()
    }
    return render(request, 'inventory/inventory.html', context)


@login_required(login_url='login')
def stock_update_view(request, id):
    """Gère la mise à jour du stock d'un produit."""
    if request.method == 'POST':
        quantity = int(request.POST['quantity'])
        supplied = bool(request.GET['supplied'] == 'true')
        record_id = request.GET['record_id'] if not supplied else None
        product = Product.objects.get(pk=id)

        if supplied:
            # Créer un nouvel enregistrement d'inventaire
            inventory_record = InventoryRecord.objects.create(
                quantity_recorded=quantity,
                done_by=request.user
            )
            inventory_record.save()
            product.inventory_records.add(inventory_record)
            product.quantity = product.quantity + quantity
            product.save()
        else:
            # Mettre à jour un enregistrement d'inventaire existant
            inventory_record = InventoryRecord.objects.get(pk=record_id)
            product.quantity += quantity - inventory_record.quantity_recorded
            inventory_record.quantity_recorded = quantity
            inventory_record.save()
            product.save()

        # Envoyer une notification de mise à jour de l'inventaire
        notification = InventoryUpdateNotification.objects.create(product=product)
        for user in User.objects.all(): notification.target.add(user)
        notification.save()

        # Gérer les notifications de stock faible
        stock_notifications = LowStockNotification.objects.filter(product=product, type=NotificationTypeChoices.low_stock)
        if len(stock_notifications) > 0:
            for notification in stock_notifications:
                if product.quantity < product.min_stock:
                    notification.save()
                else:
                    notification.delete()
        elif product.quantity < product.min_stock:
            notification = LowStockNotification.objects.create(product=product)
            for user in User.objects.all(): notification.target.add(user)
            notification.save()

        return redirect('inventory')